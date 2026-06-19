import subprocess
import tempfile
import unittest
import os
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PYTHON = Path(sys.executable)
SCRIPT = ROOT / "scripts" / "build_tables.py"
FIXTURE_DIR = Path(os.environ["HYSTERESIS_TEST_DATA_DIR"]) if os.environ.get("HYSTERESIS_TEST_DATA_DIR") else None
RAMP_FILE = Path(os.environ["HYSTERESIS_RAMP_FILE"]) if os.environ.get("HYSTERESIS_RAMP_FILE") else None


def read_wide_pair(path, specimen):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        target_col = None
        for col in range(2, ws.max_column + 1, 2):
            if ws.cell(row=3, column=col).value == specimen:
                target_col = col
                break
        if target_col is None:
            raise AssertionError(f"{specimen} not found in {path}")

        values = []
        for row in range(4, ws.max_row + 1):
            disp = ws.cell(row=row, column=target_col - 1).value
            val = ws.cell(row=row, column=target_col).value
            if disp is None and val is None:
                continue
            values.append((float(disp), float(val)))
        return values
    finally:
        wb.close()


class LoopTableGenerationTest(unittest.TestCase):
    def assert_pairs_close(self, actual, expected, tol=1e-8):
        self.assertEqual(len(actual), len(expected))
        for (ad, av), (ed, ev) in zip(actual, expected):
            self.assertLessEqual(abs(ad - ed), tol)
            self.assertLessEqual(abs(av - ev), max(tol, abs(ev) * 1e-10))

    def run_build_tables(self, out_dir, specimens):
        if FIXTURE_DIR is None or RAMP_FILE is None:
            self.skipTest("Set HYSTERESIS_TEST_DATA_DIR and HYSTERESIS_RAMP_FILE to run fixture tests.")
        subprocess.run(
            [
                str(PYTHON),
                "-X",
                "utf8",
                str(SCRIPT),
                "--energy-dir",
                str(FIXTURE_DIR),
                "--ramp-file",
                str(RAMP_FILE),
                "--output-dir",
                str(out_dir),
                "--specimens",
                ",".join(specimens),
            ],
            cwd=str(ROOT),
            check=True,
        )

    def test_build_tables_accepts_explicit_dirs_and_matches_reference_energy_stiffness(self):
        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            self.run_build_tables(out_dir, ["SHELL-1", "SHELL-2", "SHELL-3"])

            for specimen in ["SHELL-1", "SHELL-2", "SHELL-3"]:
                actual_energy = read_wide_pair(out_dir / "skeleton_energy.xlsx", specimen)
                expected_energy = read_wide_pair(FIXTURE_DIR / "energy.xlsx", specimen)
                self.assert_pairs_close(actual_energy, expected_energy)

                actual_stiffness = read_wide_pair(out_dir / "skeleton_stiffness.xlsx", specimen)
                expected_stiffness = read_wide_pair(FIXTURE_DIR / "stiffness.xlsx", specimen)
                self.assert_pairs_close(actual_stiffness, expected_stiffness)

    def test_loop_details_aligns_loopno_to_displacement(self):
        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            self.run_build_tables(out_dir, ["SHELL-1"])

            wb = load_workbook(out_dir / "loop_details.xlsx", data_only=True, read_only=True)
            try:
                ws = wb["SHELL-1"]
                self.assertEqual(
                    [ws.cell(row=1, column=i).value for i in range(1, 5)],
                    ["LoopNo", "Displacement", "1stHalfEnergy", "2stHalfEnergy"],
                )
                self.assertEqual(ws.cell(row=2, column=1).value, 1)
                self.assertEqual(ws.cell(row=2, column=2).value, 6)
                self.assertEqual(ws.cell(row=3, column=1).value, 2)
                self.assertEqual(ws.cell(row=3, column=2).value, 7.5)
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
