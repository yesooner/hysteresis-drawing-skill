# -*- coding: utf-8 -*-
"""
Build project tables from *_LoopAnalysisResults.csv and ramp.txt.

Outputs:
- skeleton_energy.xlsx: accumulated loop energy versus displacement.
- skeleton_stiffness.xlsx: average secant stiffness versus displacement.
- loop_details.xlsx: full loop audit table with LoopNo aligned to displacement.
"""

import csv
import os
import argparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


SPECIMEN_ORDER = [
    "RC",
    "BZ",
    "KJGJ18",
    "SHELL-1",
    "SHELL-2",
    "SHELL-3",
    "SHELL-3-NONE",
    "SHELL-3-RC",
]

DISPLAY_NAMES = {
    "RC": "RC",
    "BZ": "UHPC-BZ",
    "KJGJ18": "UHPC-KJGJ18",
    "SHELL-1": "SHELL-1",
    "SHELL-2": "SHELL-2",
    "SHELL-3": "SHELL-3",
    "SHELL-3-NONE": "SHELL-3-NONE",
    "SHELL-3-RC": "SHELL-3-RC",
}

DETAIL_FIELDS = [
    "LoopNo",
    "Displacement",
    "1stHalfEnergy",
    "2stHalfEnergy",
    "LoopEnergy",
    "EnergyRatio[%]",
    "AccumulatedEnergyRatio[%]",
    "1stMaxDisp",
    "1stMaxDispForce",
    "2ndMaxDisp",
    "2ndMaxDispForce",
    "1stEquivalentViscousDampRatio[%]",
    "2ndEquivalentViscousDampRatio[%]",
    "1stSecantStiffness",
    "2ndSecantStiffness",
]

CSV_FIELD_MAP = {
    "LoopNo": "LoopNo.",
    "1stHalfEnergy": "1stHalfEnergy",
    "2stHalfEnergy": "2stHalfEnergy",
    "LoopEnergy": "LoopEnergy",
    "EnergyRatio[%]": "EnergyRatio[%]",
    "AccumulatedEnergyRatio[%]": "AccumulatedEnergyRatio[%]",
    "1stMaxDisp": "1stMaxDisp",
    "1stMaxDispForce": "1stMaxDispForce",
    "2ndMaxDisp": "2ndMaxDisp",
    "2ndMaxDispForce": "2ndMaxDispForce",
    "1stEquivalentViscousDampRatio[%]": "1stEquivalentViscousDampRatio[%]",
    "2ndEquivalentViscousDampRatio[%]": "2ndEquivalentViscousDampRatio[%]",
    "1stSecantStiffness": "1stSecantStiffness",
    "2ndSecantStiffness": "2ndSecantStiffness",
}


def to_number(value):
    if value is None or value == "":
        return None
    try:
        value = float(value)
    except ValueError:
        return value
    if value.is_integer():
        return int(value)
    return value


def read_ramp(filepath):
    """Return {LoopNo: nominal positive displacement} from ramp.txt."""
    disp_map = {}
    with open(filepath, "r", encoding="utf-8") as f:
        rows = [line.strip().split() for line in f if line.strip()]

    loop_no = 1
    i = 1  # skip initial 0,0
    while i < len(rows):
        if len(rows[i]) >= 2:
            disp = abs(float(rows[i][1]))
            if disp > 0:
                disp_map[loop_no] = disp
        loop_no += 1
        i += 2
    return disp_map


def read_loop_csv(filepath, disp_map):
    """Read one LoopAnalysisResults.csv and align each LoopNo to displacement."""
    rows = []
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            try:
                loop_no = int(float(raw["LoopNo."]))
            except (KeyError, TypeError, ValueError):
                continue

            item = {
                "LoopNo": loop_no,
                "Displacement": disp_map.get(loop_no),
            }
            for out_field, csv_field in CSV_FIELD_MAP.items():
                if out_field == "LoopNo":
                    continue
                item[out_field] = to_number(raw.get(csv_field))

            s1 = item.get("1stSecantStiffness")
            s2 = item.get("2ndSecantStiffness")
            item["AvgSecantStiffness"] = (
                (float(s1) + float(s2)) / 2.0
                if isinstance(s1, (int, float)) and isinstance(s2, (int, float))
                else None
            )
            rows.append(item)
    return rows


def build_accumulated_energy(rows):
    accumulated = 0.0
    pairs = []
    for row in rows:
        loop_energy = row.get("LoopEnergy")
        disp = row.get("Displacement")
        if not isinstance(loop_energy, (int, float)) or disp is None:
            continue
        accumulated += float(loop_energy)
        pairs.append((disp, accumulated))
    return pairs


def parse_args():
    parser = argparse.ArgumentParser(description="Build loop energy/stiffness tables.")
    parser.add_argument("--energy-dir", required=True, help="Directory containing *_LoopAnalysisResults.csv.")
    parser.add_argument("--ramp-file", required=True, help="Path to ramp.txt.")
    parser.add_argument("--output-dir", required=True, help="Directory for generated xlsx files.")
    parser.add_argument(
        "--specimens",
        required=True,
        help="Comma-separated specimen names to process, matching CSV prefixes.",
    )
    return parser.parse_args()


def parse_specimens(specimens_text):
    return [item.strip() for item in specimens_text.split(",") if item.strip()]


def create_wide_workbook(value_title, value_unit, data_rows, specimen_order):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")

    for i, spec in enumerate(specimen_order):
        col_disp = 1 + i * 2
        col_val = 2 + i * 2

        ws.cell(row=1, column=col_disp, value="位移").font = header_font
        ws.cell(row=1, column=col_val, value=value_title).font = header_font
        ws.cell(row=2, column=col_disp, value="mm").font = header_font
        ws.cell(row=2, column=col_val, value=value_unit).font = header_font
        ws.cell(row=3, column=col_val, value=DISPLAY_NAMES.get(spec, spec)).font = header_font

        for r in (1, 2, 3):
            ws.cell(row=r, column=col_disp).alignment = center
            ws.cell(row=r, column=col_val).alignment = center

        for row_idx, (disp, value) in enumerate(data_rows.get(spec, []), start=4):
            ws.cell(row=row_idx, column=col_disp, value=disp)
            ws.cell(row=row_idx, column=col_val, value=value)

    for col in range(1, 1 + len(specimen_order) * 2):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16
    return wb


def create_loop_detail_workbook(detail_data, specimen_order):
    """Create one sheet per specimen with displacement and LoopNo aligned."""
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    header_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")

    for spec in specimen_order:
        rows = detail_data.get(spec)
        if not rows:
            continue

        ws = wb.create_sheet(title=DISPLAY_NAMES.get(spec, spec)[:31])
        for col_idx, field in enumerate(DETAIL_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.alignment = center

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, field in enumerate(DETAIL_FIELDS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(field))

        for col_idx in range(1, len(DETAIL_FIELDS) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

    return wb


def main():
    args = parse_args()
    specimen_order = parse_specimens(args.specimens)
    os.makedirs(args.output_dir, exist_ok=True)

    disp_map = read_ramp(args.ramp_file)
    print("Ramp displacement levels:")
    for loop_no, disp in sorted(disp_map.items()):
        print(f"  Loop {loop_no}: {disp} mm")

    detail_data = {}
    energy_data = {}
    stiffness_data = {}

    for spec in specimen_order:
        csv_path = os.path.join(args.energy_dir, f"{spec}_LoopAnalysisResults.csv")
        if not os.path.exists(csv_path):
            print(f"  [WARN] missing {csv_path}")
            continue

        rows = read_loop_csv(csv_path, disp_map)
        detail_data[spec] = rows
        energy_data[spec] = build_accumulated_energy(rows)
        stiffness_data[spec] = [
            (row["Displacement"], row["AvgSecantStiffness"])
            for row in rows
            if row.get("Displacement") is not None and row.get("AvgSecantStiffness") is not None
        ]
        print(f"  {spec}: {len(rows)} loops")

    wb_energy = create_wide_workbook("耗能", "kN·mm", energy_data, specimen_order)
    out_energy = os.path.join(args.output_dir, "skeleton_energy.xlsx")
    wb_energy.save(out_energy)

    wb_stiffness = create_wide_workbook("刚度", "kN/mm", stiffness_data, specimen_order)
    out_stiffness = os.path.join(args.output_dir, "skeleton_stiffness.xlsx")
    wb_stiffness.save(out_stiffness)

    wb_detail = create_loop_detail_workbook(detail_data, specimen_order)
    out_detail = os.path.join(args.output_dir, "loop_details.xlsx")
    wb_detail.save(out_detail)

    print(f"Saved: {out_energy}")
    print(f"Saved: {out_stiffness}")
    print(f"Saved: {out_detail}")


if __name__ == "__main__":
    main()
