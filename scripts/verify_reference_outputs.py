# -*- coding: utf-8 -*-
"""
Verify skill-generated tables against reference baseline files.

This script intentionally has no default input/output directories. The caller
must provide every path to avoid mixing fixture, project, and production data.
"""

import argparse
import csv
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
PYTHON = Path(sys.executable)
BUILD_TABLES = Path(__file__).with_name("build_tables.py")

CSV_FIELD_MAP = {
    "LoopNo": "LoopNo.",
    "1stHalfEnergy": "1stHalfEnergy",
    "2stHalfEnergy": "2stHalfEnergy",
    "LoopEnergy": "LoopEnergy",
    "EnergyRatio[%]": "EnergyRatio[%]",
    "AccumulatedEnergyRatio[%]": "AccumulatedEnergyRatio[%]",
    "1stMaxDisp": "1stMaxDisp",
    "1stMaxDispForce": "1stMaxDispForce",
    "2ndMaxDisp": "2ndMaxDisp",
    "2ndMaxDispForce": "2ndMaxDispForce",
    "1stEquivalentViscousDampRatio[%]": "1stEquivalentViscousDampRatio[%]",
    "2ndEquivalentViscousDampRatio[%]": "2ndEquivalentViscousDampRatio[%]",
    "1stSecantStiffness": "1stSecantStiffness",
    "2ndSecantStiffness": "2ndSecantStiffness",
}


def parse_args():
    parser = argparse.ArgumentParser(description="Verify hysteresis skill outputs against reference baselines.")
    parser.add_argument("--raw-hysteresis", required=True, help="Reference raw hysteresis .xls file.")
    parser.add_argument("--loop-dir", required=True, help="Directory containing *_LoopAnalysisResults.csv.")
    parser.add_argument("--ramp-file", required=True, help="Ramp file used to map LoopNo to displacement.")
    parser.add_argument("--skeleton-ref", required=True, help="Reference skeleton.xlsx baseline.")
    parser.add_argument("--stiffness-ref", required=True, help="Reference stiffness.xlsx baseline.")
    parser.add_argument("--output-dir", required=True, help="Directory for generated outputs and reports.")
    parser.add_argument("--specimens", required=True, help="Comma-separated specimens to verify.")
    return parser.parse_args()


def parse_specimens(text):
    return [item.strip() for item in text.split(",") if item.strip()]


def read_ramp(path):
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            parts = line.strip().split()
            if len(parts) >= 2:
                rows.append(float(parts[1]))
    result = {}
    loop_no = 1
    i = 1
    while i < len(rows):
        if abs(rows[i]) > 0:
            result[loop_no] = abs(rows[i])
        loop_no += 1
        i += 2
    return result


def read_wide_pair(path, specimen):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        target_col = None
        for col in range(2, ws.max_column + 1, 2):
            if ws.cell(row=3, column=col).value == specimen:
                target_col = col
                break
        if target_col is None:
            raise ValueError(f"{specimen} not found in {path}")

        values = []
        for row in range(4, ws.max_row + 1):
            disp = ws.cell(row=row, column=target_col - 1).value
            value = ws.cell(row=row, column=target_col).value
            if disp is None and value is None:
                continue
            values.append((float(disp), float(value)))
        return values
    finally:
        wb.close()


def max_pair_diff(actual, expected):
    count = min(len(actual), len(expected))
    max_disp = 0.0
    max_val = 0.0
    for i in range(count):
        max_disp = max(max_disp, abs(actual[i][0] - expected[i][0]))
        max_val = max(max_val, abs(actual[i][1] - expected[i][1]))
    return {
        "rows_actual": len(actual),
        "rows_expected": len(expected),
        "max_disp_diff": max_disp,
        "max_value_diff": max_val,
    }


def read_csv_rows(path):
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def as_float(value):
    if value is None or value == "":
        return None
    return float(value)


def compare_loop_details(output_dir, loop_dir, ramp_map, specimens):
    details = []
    wb = load_workbook(Path(output_dir) / "loop_details.xlsx", data_only=True, read_only=True)
    try:
        for spec in specimens:
            ws = wb[spec]
            headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
            header_col = {name: i + 1 for i, name in enumerate(headers)}
            csv_rows = read_csv_rows(Path(loop_dir) / f"{spec}_LoopAnalysisResults.csv")
            max_diff = 0.0
            mismatch_count = 0

            for idx, csv_row in enumerate(csv_rows, start=2):
                loop_no = int(float(csv_row["LoopNo."]))
                disp = ws.cell(row=idx, column=header_col["Displacement"]).value
                max_diff = max(max_diff, abs(float(disp) - float(ramp_map[loop_no])))
                for out_field, csv_field in CSV_FIELD_MAP.items():
                    actual = ws.cell(row=idx, column=header_col[out_field]).value
                    expected = as_float(csv_row[csv_field])
                    if out_field == "LoopNo":
                        expected = loop_no
                    if expected is None and actual is None:
                        continue
                    diff = abs(float(actual) - float(expected))
                    max_diff = max(max_diff, diff)
                    if diff > max(1e-8, abs(float(expected)) * 1e-10):
                        mismatch_count += 1

            details.append({
                "table": "loop_details",
                "specimen": spec,
                "rows_actual": ws.max_row - 1,
                "rows_expected": len(csv_rows),
                "max_diff": max_diff,
                "mismatch_count": mismatch_count,
            })
    finally:
        wb.close()
    return details


def read_raw_series(raw_hysteresis):
    df = pd.read_excel(raw_hysteresis, header=None, engine="xlrd")
    series = {}
    for col in range(1, df.shape[1], 2):
        name = df.iloc[2, col]
        if not isinstance(name, str):
            continue
        x = pd.to_numeric(df.iloc[3:, col - 1], errors="coerce")
        y = pd.to_numeric(df.iloc[3:, col], errors="coerce")
        valid = x.notna() & y.notna()
        series[name] = list(zip(x[valid].astype(float), y[valid].astype(float)))
    return series


def read_skeleton_series(skeleton_ref):
    wb = load_workbook(skeleton_ref, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        series = {}
        for col in range(2, ws.max_column + 1, 2):
            name = ws.cell(row=3, column=col).value
            if not isinstance(name, str):
                continue
            points = []
            for row in range(4, ws.max_row + 1):
                x = ws.cell(row=row, column=col - 1).value
                y = ws.cell(row=row, column=col).value
                if x is None and y is None:
                    continue
                points.append((float(x), float(y)))
            series[name] = points
        return series
    finally:
        wb.close()


def verify_skeleton_points(raw_hysteresis, skeleton_ref, specimens):
    raw = read_raw_series(raw_hysteresis)
    skeleton = read_skeleton_series(skeleton_ref)
    details = []
    for spec in specimens:
        raw_points = raw.get(spec, [])
        raw_set = {(round(x, 8), round(y, 8)) for x, y in raw_points}
        missing = 0
        max_nearest_dx = 0.0
        max_nearest_dy = 0.0
        for sx, sy in skeleton.get(spec, []):
            if (round(sx, 8), round(sy, 8)) in raw_set:
                continue
            missing += 1
            if raw_points:
                nx, ny = min(raw_points, key=lambda p: abs(p[0] - sx) + abs(p[1] - sy) / 1000.0)
                max_nearest_dx = max(max_nearest_dx, abs(nx - sx))
                max_nearest_dy = max(max_nearest_dy, abs(ny - sy))
        details.append({
            "table": "skeleton",
            "specimen": spec,
            "rows_actual": len(skeleton.get(spec, [])),
            "rows_expected": len(skeleton.get(spec, [])),
            "missing_from_raw": missing,
            "max_nearest_dx": max_nearest_dx,
            "max_nearest_dy": max_nearest_dy,
        })
    return details


def write_reports(output_dir, rows):
    output_dir = Path(output_dir)
    csv_path = output_dir / "verification_details.csv"
    keys = sorted({key for row in rows for key in row})
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(rows)

    txt_path = output_dir / "verification_report.txt"
    lines = ["Reference baseline verification report", ""]
    for row in rows:
        lines.append(", ".join(f"{key}={row.get(key)}" for key in keys if key in row))
    txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return txt_path, csv_path


def main():
    args = parse_args()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    specimens = parse_specimens(args.specimens)

    subprocess.run(
        [
            str(PYTHON),
            "-X",
            "utf8",
            str(BUILD_TABLES),
            "--energy-dir",
            args.loop_dir,
            "--ramp-file",
            args.ramp_file,
            "--output-dir",
            str(output_dir),
            "--specimens",
            ",".join(specimens),
        ],
        cwd=str(ROOT),
        check=True,
    )

    rows = []
    generated_stiffness = output_dir / "skeleton_stiffness.xlsx"
    for spec in specimens:
        diff = max_pair_diff(
            read_wide_pair(generated_stiffness, spec),
            read_wide_pair(args.stiffness_ref, spec),
        )
        diff.update({"table": "stiffness", "specimen": spec})
        rows.append(diff)

    ramp_map = read_ramp(args.ramp_file)
    rows.extend(compare_loop_details(output_dir, args.loop_dir, ramp_map, specimens))
    rows.extend(verify_skeleton_points(args.raw_hysteresis, args.skeleton_ref, specimens))

    txt_path, csv_path = write_reports(output_dir, rows)
    print(f"Saved: {txt_path}")
    print(f"Saved: {csv_path}")


if __name__ == "__main__":
    main()
