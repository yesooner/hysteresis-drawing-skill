# -*- coding: utf-8 -*-
"""
Draw plots from generated hysteresis postprocess workbooks.

All paths are explicit command-line arguments. No input/output directories are
hard-coded in this script.
"""

import argparse
from pathlib import Path

from openpyxl import load_workbook


def parse_args():
    parser = argparse.ArgumentParser(description="Draw hysteresis postprocess plots.")
    parser.add_argument("--input-dir", required=True, help="Directory containing generated xlsx files.")
    parser.add_argument("--plot-dir", required=True, help="Directory for generated plot files.")
    parser.add_argument(
        "--tables",
        required=True,
        help="Comma-separated workbook names, e.g. skeleton_energy.xlsx,skeleton_stiffness.xlsx,loop_details.xlsx.",
    )
    parser.add_argument("--format", default="png", choices=["png", "pdf", "svg"], help="Plot file format.")
    return parser.parse_args()


def sanitize_filename(text):
    keep = []
    for char in str(text):
        if char.isalnum() or char in ("-", "_", "."):
            keep.append(char)
        else:
            keep.append("_")
    return "".join(keep).strip("_") or "plot"


def read_wide_pairs(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        series = {}
        for col in range(2, ws.max_column + 1, 2):
            name = ws.cell(row=3, column=col).value
            if not name:
                continue
            points = []
            for row in range(4, ws.max_row + 1):
                x = ws.cell(row=row, column=col - 1).value
                y = ws.cell(row=row, column=col).value
                if x is None or y is None:
                    continue
                points.append((float(x), float(y)))
            if points:
                series[str(name)] = points
        return series
    finally:
        wb.close()


def read_loop_detail(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        series = {}
        for ws in wb.worksheets:
            headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            header_col = {name: idx + 1 for idx, name in enumerate(headers)}
            if "Displacement" not in header_col:
                continue
            for metric in ("LoopEnergy", "1stSecantStiffness", "2ndSecantStiffness"):
                if metric not in header_col:
                    continue
                points = []
                for row in range(2, ws.max_row + 1):
                    x = ws.cell(row=row, column=header_col["Displacement"]).value
                    y = ws.cell(row=row, column=header_col[metric]).value
                    if x is None or y is None:
                        continue
                    points.append((float(x), float(y)))
                if points:
                    series[f"{ws.title}_{metric}"] = points
        return series
    finally:
        wb.close()


def draw_series(series, title, xlabel, ylabel, out_path):
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(8.0, 5.0), dpi=160)
    for name, points in series.items():
        xs = [p[0] for p in points]
        ys = [p[1] for p in points]
        ax.plot(xs, ys, marker="o", linewidth=1.2, markersize=3.2, label=name)
    ax.set_title(title)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.grid(True, linestyle="--", linewidth=0.5, alpha=0.55)
    ax.legend(fontsize=8)
    fig.tight_layout()
    fig.savefig(out_path)
    plt.close(fig)


def infer_axis_labels(filename):
    lower = filename.lower()
    if "energy" in lower:
        return "Displacement (mm)", "Energy (kN·mm)"
    if "stiffness" in lower:
        return "Displacement (mm)", "Secant stiffness (kN/mm)"
    if "skeleton" in lower or "combined" in lower:
        return "Displacement (mm)", "Load (kN)"
    return "Displacement", "Value"


def main():
    args = parse_args()
    input_dir = Path(args.input_dir)
    plot_dir = Path(args.plot_dir)
    plot_dir.mkdir(parents=True, exist_ok=True)

    for table in [item.strip() for item in args.tables.split(",") if item.strip()]:
        path = input_dir / table
        if not path.exists():
            raise FileNotFoundError(path)

        if path.name == "loop_details.xlsx":
            series = read_loop_detail(path)
        else:
            series = read_wide_pairs(path)

        xlabel, ylabel = infer_axis_labels(path.name)
        out_path = plot_dir / f"{sanitize_filename(path.stem)}.{args.format}"
        draw_series(series, path.stem, xlabel, ylabel, out_path)
        print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
